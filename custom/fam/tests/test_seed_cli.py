"""Smoke tests for scripts/data_roundtrip.py (Task 6).

Imports the script's main([...]) directly against a tmp DB + tmp snapshot
dir -- conftest's autouse _isolate_db fixture already points FAM_DB at a
tmp path, so the CLI (which resolves --db first, FAM_DB second) picks it
up with no extra plumbing needed, same as any other fam.cli test.
"""
import json
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

SCRIPTS_DIR = Path(__file__).resolve().parent.parent / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))

import data_roundtrip  # noqa: E402


@pytest.fixture()
def seeded_db(db):
    """A tiny live DB with one place + one event referencing it, via the
    real fam domain modules (mirrors how seed.export_rows expects data to
    look), committed so the CLI's own famdb.connect() sees it."""
    from fam import cal, places

    pl = places.add(db, "Дом", address="ул. Тестовая, 1")
    cal.add(db, "Плановое событие", "2099-01-01T10:00:00+00:00",
            place="Дом", transport="car")
    db.commit()
    return db


def test_export_creates_xlsx_and_snapshot(seeded_db, tmp_path, monkeypatch):
    monkeypatch.setattr(data_roundtrip, "DEFAULT_SNAPSHOT_DIR", tmp_path / "seeding")
    out = tmp_path / "data.xlsx"
    rc = data_roundtrip.main(["export", "--out", str(out)])
    assert rc == 0

    assert out.exists()
    wb = load_workbook(out)
    assert "Места" in wb.sheetnames
    assert "README" in wb.sheetnames

    snap_dir = tmp_path / "seeding"
    snaps = list(snap_dir.glob("export-*.json"))
    assert len(snaps) == 1
    snap = json.loads(snaps[0].read_text(encoding="utf-8"))
    assert "sheets" in snap
    assert "Места" in snap["sheets"]


def test_diff_on_untouched_export_is_clean(seeded_db, tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(data_roundtrip, "DEFAULT_SNAPSHOT_DIR", tmp_path / "seeding")
    out = tmp_path / "data.xlsx"
    assert data_roundtrip.main(["export", "--out", str(out)]) == 0
    snap_path = next((tmp_path / "seeding").glob("export-*.json"))

    capsys.readouterr()
    rc = data_roundtrip.main(["diff", "--file", str(out), "--snapshot", str(snap_path)])
    captured = capsys.readouterr()

    assert rc == 0
    assert "изменений нет" in captured.out.casefold()


def test_diff_on_corrupted_transport_reports_conflict(seeded_db, tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(data_roundtrip, "DEFAULT_SNAPSHOT_DIR", tmp_path / "seeding")
    out = tmp_path / "data.xlsx"
    assert data_roundtrip.main(["export", "--out", str(out)]) == 0
    snap_path = next((tmp_path / "seeding").glob("export-*.json"))

    wb = load_workbook(out)
    ws = wb["События"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = headers.index("транспорт") + 1
    ws.cell(row=2, column=col, value="на ковре-самолёте")
    wb.save(out)

    capsys.readouterr()
    rc = data_roundtrip.main(["diff", "--file", str(out), "--snapshot", str(snap_path)])
    captured = capsys.readouterr()

    assert rc == 2
    assert "⚠" in captured.out
