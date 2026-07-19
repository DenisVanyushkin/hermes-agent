import openpyxl
from fam import seed, seed_xlsx


def test_workbook_roundtrip(tmp_path):
    rows = {"Планы": [{"id": 1, "title": "Пироги", "deadline": "2026-08-01",
                       "place": None, "person": None, "notes": None, "link": None}], **{
            s: [] for s in seed.SHEETS if s != "Планы"}}
    path = tmp_path / "data.xlsx"
    seed_xlsx.write_workbook(rows, path)
    back = seed_xlsx.read_workbook(path)
    assert seed.normalize_row("Планы", back["Планы"][0]) == seed.normalize_row(
        "Планы", {"id": 1, "название": "Пироги", "срок": "2026-08-01"})


def test_workbook_has_comments_and_readme(tmp_path):
    seed_xlsx.write_workbook({s: [] for s in seed.SHEETS}, tmp_path / "d.xlsx")
    wb = openpyxl.load_workbook(tmp_path / "d.xlsx")
    assert wb.sheetnames[0] == "README"
    ws = wb["События"]
    assert all(c.comment and c.comment.text.strip() for c in ws[1])   # комментарий у каждого заголовка
    dvs = list(ws.data_validations.dataValidation)                     # dropdown транспорта
    assert any('"машина,пешком,общественный"' == dv.formula1 for dv in dvs)


def test_read_workbook_skips_empty_rows(tmp_path):
    rows = {s: [] for s in seed.SHEETS}
    rows["Места"] = [
        {"id": 1, "name": "Дом", "address": None, "category": None, "lat": None,
         "lon": None, "gis_url": None, "travel_min": None, "aliases": [], "notes": None},
    ]
    path = tmp_path / "e.xlsx"
    seed_xlsx.write_workbook(rows, path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Места"]
    ws.append([None] * ws.max_column)  # trailing blank row
    wb.save(path)
    back = seed_xlsx.read_workbook(path)
    assert len(back["Места"]) == 1


def test_readonly_and_id_columns_are_greyed(tmp_path):
    seed_xlsx.write_workbook({s: [] for s in seed.SHEETS}, tmp_path / "g.xlsx")
    wb = openpyxl.load_workbook(tmp_path / "g.xlsx")
    ws = wb["Планы"]
    spec = seed.SHEETS["Планы"]
    idx = {col.header: i + 1 for i, col in enumerate(spec.columns)}
    link_col = idx["связь с событием"]
    cell = ws.cell(row=1, column=link_col)
    assert cell.fill.fgColor.rgb in ("00DDDDDD", "FFDDDDDD")
