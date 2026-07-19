"""xlsx-слой: workbook writer/reader поверх словаря листов/колонок seed.py.

write_workbook(rows_by_sheet, path) пишет .xlsx: README первым листом,
дальше по одному листу на seed.SHEETS с bold-заголовками, комментарием на
каждом заголовке (Col.comment), серой заливкой id/readonly-колонок,
dropdown-валидацией для доменных полей и замороженной первой строкой.

read_workbook(path) -> {sheet: [row_dict, ...]} читает обратно, ключи
row_dict — русские заголовки (col.header), как ждёт seed.normalize_row.
"""
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from . import seed

_GREY = PatternFill("solid", fgColor="DDDDDD")
_BOLD = Font(bold=True)
_AUTHOR = "Гермес"

# header -> список опций dropdown'а (по колонкам, где это применимо)
_DROPDOWNS = {
    "транспорт": "машина,пешком,общественный",
    "тип": "человек,группа",
    "категория": "продукты,аптека",
    "включено": "да,нет",
}

_README_TEXT = [
    "Этот файл — источник истины: что здесь написано, то и будет в базе Гермеса "
    "после импорта.",
    "",
    "Удалил строку — удалил запись. Новая строка без id — новая запись.",
    "Колонку «id» не трогать и не копировать между строками: по ней Гермес "
    "узнаёт, что редактируется, а не создаётся заново.",
    "",
    "Дата и время — в формате ГГГГ-ММ-ДД ЧЧ:ММ, время алматинское.",
    "Для места вместо lat/lon можно вставить 2ГИС-ссылку (в т.ч. короткую "
    "go.2gis.com/…) — координаты подтянутся при импорте сами.",
    "Колонка «участники» заполняется именами людей, не группами (группы "
    "разворачиваются в имена участников при экспорте).",
    "",
    "Серые колонки — только для чтения (заполняет сам Гермес), при импорте "
    "игнорируются.",
    "",
    "ВНИМАНИЕ: колонка «slug» на листе «Люди» завязана на правила напоминаний "
    "(например, ранние стадии сборов для Таи). Меняй её только осознанно.",
]


def _canon_row_to_cells(sheet, row):
    """Каноническая строка (col.key -> canon value) -> dict header -> ячейка,
    применяя Col.to_xlsx там, где он задан (canon -> русская метка/CSV)."""
    spec = seed.SHEETS[sheet]
    out = {}
    for col in spec.columns:
        v = row.get(col.key)
        if v is not None and col.to_xlsx:
            v = col.to_xlsx(v)
        out[col.header] = v
    return out


def _write_readme(wb):
    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 100
    for i, line in enumerate(_README_TEXT, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=13)


def _write_sheet(wb, sheet, rows):
    spec = seed.SHEETS[sheet]
    ws = wb.create_sheet(sheet)
    for c_idx, col in enumerate(spec.columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col.header)
        cell.font = _BOLD
        cell.comment = Comment(col.comment, _AUTHOR)
        if col.key == "id" or col.readonly:
            cell.fill = _GREY
        ws.column_dimensions[cell.column_letter].width = max(len(col.header), 18)

        options = _DROPDOWNS.get(col.header)
        if options:
            dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{cell.column_letter}2:{cell.column_letter}500")

    for r_idx, row in enumerate(rows, start=2):
        cells = _canon_row_to_cells(sheet, row)
        for c_idx, col in enumerate(spec.columns, start=1):
            v = cells.get(col.header)
            ws.cell(row=r_idx, column=c_idx, value=v)
            if col.key == "id" or col.readonly:
                ws.cell(row=r_idx, column=c_idx).fill = _GREY

    ws.freeze_panes = "A2"


def write_workbook(rows_by_sheet, path):
    wb = Workbook()
    wb.remove(wb.active)  # дефолтный "Sheet"
    _write_readme(wb)
    for sheet in seed.SHEETS:
        _write_sheet(wb, sheet, rows_by_sheet.get(sheet, []))
    wb.save(path)


def read_workbook(path):
    wb = load_workbook(path, data_only=True)
    out = {}
    for sheet, spec in seed.SHEETS.items():
        if sheet not in wb.sheetnames:
            out[sheet] = []
            continue
        ws = wb[sheet]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in raw):
                continue
            rows.append({headers[i]: raw[i] for i in range(len(headers))})
        out[sheet] = rows
    return out
